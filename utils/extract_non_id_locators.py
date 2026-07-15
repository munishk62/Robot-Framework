#!/usr/bin/env python3
"""
Extract all locators from page files that don't use ID selectors.
Creates an Excel file with the results.
"""

import os
import re
from pathlib import Path
from typing import List, Dict, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def is_id_locator(locator_value: str) -> bool:
    """
    Check if a locator uses ID selector.

    Args:
        locator_value: The XPath or CSS selector string

    Returns:
        True if it uses ID, False otherwise
    """
    if not locator_value or not isinstance(locator_value, str):
        return False

    # Strip quotes and whitespace
    locator = locator_value.strip().strip('"').strip("'")

    # Check for CSS ID selector
    if locator.startswith('#') and ' ' not in locator.split('#')[1].split('[')[0]:
        return True

    # Check for XPath with @id
    if '//*[@id=' in locator or '//[@id=' in locator:
        # Make sure it's a direct @id attribute, not contains(@id or other functions
        if '@id=' in locator and 'contains(@id' not in locator and '@id!=' not in locator:
            return True

    # Check for id() function in XPath
    if locator.startswith('id('):
        return True

    return False

def extract_locators_from_file(file_path: Path) -> List[Tuple[str, str, str]]:
    """
    Extract all locators from a Python page file.

    Args:
        file_path: Path to the Python file

    Returns:
        List of tuples (locator_name, locator_value, file_name)
    """
    locators = []

    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        i = 0
        while i < len(lines):
            line = lines[i].strip()

            # Skip comments and empty lines
            if not line or line.startswith('#'):
                i += 1
                continue

            # Check if this is a locator assignment (UPPERCASE_NAME = ...)
            if '=' in line:
                parts = line.split('=', 1)
                if len(parts) == 2:
                    var_name = parts[0].strip()
                    # Must be all uppercase with underscores
                    if var_name and var_name.replace('_', '').isupper() and var_name[0].isupper():
                        value_part = parts[1].strip()
                        locator_value = extract_string_value(lines, i, value_part)

                        if locator_value and not is_id_locator(locator_value):
                            locators.append((var_name, locator_value, file_path.name))

            i += 1

    except Exception as e:
        print(f"Error processing {file_path}: {e}")

    return locators

def extract_string_value(lines: List[str], start_index: int, value_part: str) -> str:
    """
    Extract complete string value from Python assignment.
    Handles: "string", 'string', ("string"), ('string'), '''string''', \"\"\"string\"\"\"

    Args:
        lines: List of all lines in the file
        start_index: Index of the line where the assignment starts
        value_part: The part after the = sign

    Returns:
        Complete string value
    """
    value_part = value_part.strip()

    # Handle opening parenthesis
    if value_part.startswith('('):
        value_part = value_part[1:].strip()

    # Determine quote type
    quote_type = None
    is_triple = False

    if value_part.startswith('"""') or value_part.startswith("'''"):
        is_triple = True
        quote_type = value_part[0:3]
        value_part = value_part[3:]
    elif value_part.startswith('"'):
        quote_type = '"'
        value_part = value_part[1:]
    elif value_part.startswith("'"):
        quote_type = "'"
        value_part = value_part[1:]
    else:
        return ""

    # Build the complete string
    result_parts = []
    current_text = value_part
    line_idx = start_index

    while line_idx < len(lines):
        # Look for closing quote
        if is_triple:
            if quote_type in current_text:
                pos = current_text.find(quote_type)
                result_parts.append(current_text[:pos])
                break
            else:
                result_parts.append(current_text)
        else:
            # For single quotes, find the closing quote (not escaped)
            pos = 0
            found_close = False
            while pos < len(current_text):
                if current_text[pos] == quote_type[0]:
                    # Check if it's escaped
                    if pos > 0 and current_text[pos-1] == '\\':
                        pos += 1
                        continue
                    # Found closing quote
                    result_parts.append(current_text[:pos])
                    found_close = True
                    break
                pos += 1

            if found_close:
                break
            else:
                result_parts.append(current_text)

        # Move to next line
        line_idx += 1
        if line_idx < len(lines):
            current_text = lines[line_idx].strip()
        else:
            break

    # Join and clean up
    final_value = ' '.join(result_parts).strip()

    # Remove trailing closing parenthesis if present
    if final_value.endswith(')'):
        final_value = final_value[:-1].strip()

    return final_value

def find_all_page_files(root_dir: str) -> List[Path]:
    """
    Find all *_page.py files in the directory tree.

    Args:
        root_dir: Root directory to search

    Returns:
        List of Path objects for page files
    """
    root_path = Path(root_dir)
    page_files = []

    for file_path in root_path.rglob('*_page.py'):
        page_files.append(file_path)

    return sorted(page_files)

def create_excel_report(locators_data: List[Tuple[str, str, str, str]], output_file: str):
    """
    Create an Excel file with locator information.

    Args:
        locators_data: List of tuples (file_name, module, locator_name, locator_value)
        output_file: Path to output Excel file
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Locators Without ID"

    # Define header
    headers = ['File Name', 'Module', 'Locator Name', 'Current Locator Value']
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Add data rows with text wrapping
    for row_data in locators_data:
        row_num = ws.max_row + 1
        ws.append(row_data)
        # Enable text wrapping for all cells in the row
        for cell in ws[row_num]:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        # Set row height to auto-adjust based on content
        ws.row_dimensions[row_num].height = None  # Auto-height

    # Set column widths - extra wide for locator value column
    ws.column_dimensions['A'].width = 35  # File Name
    ws.column_dimensions['B'].width = 30  # Module
    ws.column_dimensions['C'].width = 45  # Locator Name
    ws.column_dimensions['D'].width = 150  # Current Locator Value - extra wide for long XPath


    # Set row height for header
    ws.row_dimensions[1].height = 25

    # Enable filters
    ws.auto_filter.ref = ws.dimensions

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save workbook
    wb.save(output_file)
    print(f"✓ Excel report saved to: {output_file}")

def main():
    """Main execution function."""

    # Set the root directory
    root_dir = '/resources/web'

    print("=" * 70)
    print("EXTRACTING LOCATORS WITHOUT ID SELECTORS")
    print("=" * 70)

    # Find all page files
    print(f"\n1. Scanning for *_page.py files in {root_dir}...")
    page_files = find_all_page_files(root_dir)
    print(f"   Found {len(page_files)} page files")

    # Extract locators from all files
    print("\n2. Extracting locators without ID selectors...")
    all_locators = []

    for file_path in page_files:
        locators = extract_locators_from_file(file_path)

        # Get module name from path
        relative_path = file_path.relative_to(Path(root_dir))
        module = str(relative_path.parent) if str(relative_path.parent) != '.' else 'root'

        for locator_name, locator_value, file_name in locators:
            # Remove .py extension from file name
            file_name_clean = file_name.replace('.py', '') if file_name.endswith('.py') else file_name
            all_locators.append((file_name_clean, module, locator_name, locator_value))

    print(f"   Extracted {len(all_locators)} locators without ID selectors")

    # Create Excel report
    print("\n3. Creating Excel report...")
    output_file = '/locators_without_id.xlsx'
    create_excel_report(all_locators, output_file)

    # Summary by module
    print("\n4. Summary by Module:")
    print("-" * 70)

    module_counts = {}
    for _, module, _, _ in all_locators:
        module_counts[module] = module_counts.get(module, 0) + 1

    for module in sorted(module_counts.keys()):
        print(f"   {module}: {module_counts[module]} locators")

    print("\n" + "=" * 70)
    print(f"✓ COMPLETE! Total locators without ID: {len(all_locators)}")
    print(f"✓ Excel file: {output_file}")
    print("=" * 70)

if __name__ == '__main__':
    main()

