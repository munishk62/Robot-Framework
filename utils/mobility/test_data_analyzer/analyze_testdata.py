#!/usr/bin/env python3
"""
Test Data Analysis Script
Analyzes test data CSV files to track user and day usage across test cases.
Generates Excel report with PW-D format in reports/ subdirectory.

Usage:
    python3 analyze_testdata.py [CSV_FILE] [TEMPLATE_FILE] [OUTPUT_FILE]

Examples:
    # Use default files (VARIATION1.csv) - reports go to reports/ folder
    python3 analyze_testdata.py

    # Specify custom CSV file
    python3 analyze_testdata.py ../../../TestData/VARIATION2.csv

    # Specify all files
    python3 analyze_testdata.py ../../../TestData/VARIATION2.csv ../../../TestData/template_mapping_shift_data/template_week_mapping.json reports/variation2_report.xlsx
"""

import csv
import re
import json
import os
import sys
import argparse
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

# Try to import openpyxl for Excel generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Install with: pip3 install openpyxl")
    print("   Will generate CSV instead of Excel.")


class TestDataAnalyzer:
    def __init__(self, csv_file_path):
        self.csv_file_path = csv_file_path
        self.test_data = []
        self.test_case_info = defaultdict(list)  # {test_id: [rows]}
        self.user_usage_by_test = defaultdict(set)  # {test_id: set of users}
        self.user_pw_day_usage = defaultdict(set)  # {user: set of PW-D strings}
        self.date_to_pw_map = {}  # {date: PW-D string}

        # Enhanced tracking: {(user, pw_day): [test_case_ids]}
        self.user_day_test_mapping = defaultdict(
            list
        )  # Track which tests use which days for each user

    def normalize_pw_format(self, pw_string):
        """
        Normalize planning week format from PW02-D1 to PW2-D1
        Args:
            pw_string: String like 'PW02-D1' or 'PW2-D1'
        Returns:
            Normalized string like 'PW2-D1'
        """
        if not pw_string or not isinstance(pw_string, str):
            return pw_string

        # Match PW with optional leading zero + week number + D + day number
        match = re.match(r"PW0*(\d+)-D(\d+)", pw_string)
        if match:
            week = match.group(1)
            day = match.group(2)
            return f"PW{week}-D{day}"
        return pw_string

    def load_template_mapping(self, template_file_path):
        """Load template week mapping and apply to all matching users."""
        self.template_mapping = {}  # {user_id: [weeks]}
        self.user_available_days = defaultdict(set)  # {user: set of PW-D strings}

        if not os.path.exists(template_file_path):
            print(f"Warning: Template file not found: {template_file_path}")
            return

        try:
            with open(template_file_path, "r") as f:
                data = json.load(f)

            # Build manager templates
            manager_templates = {}
            for user_data in data:
                manager_id = user_data.get("id")
                weeks = user_data.get("weeks", [])
                manager_templates[manager_id] = weeks

            print(
                f"Loaded {len(manager_templates)} manager templates: {list(manager_templates.keys())}"
            )

            # Map manager templates to user patterns
            # SIT-SM-SCH-USER1 -> applies to all SIT-ESS-SCH-* users
            # SIT-SM-ALTWORK-USER1 -> applies to all SIT-ESS-ALTWORK-* or SIT-SM-ALTWORK-* users
            user_type_mapping = {
                "SIT-SM-SCH-USER1": r"SIT-(ESS|SM)-SCH-",  # All schedule users
                "SIT-SM-ALTWORK-USER1": r"SIT-(ESS|SM)-ALTWORK-",  # All altwork users
            }

            self.manager_templates = manager_templates
            self.user_type_mapping = user_type_mapping

        except Exception as e:
            print(f"Error loading template mapping: {e}")
            import traceback

            traceback.print_exc()

    def calculate_available_days(self):
        """Calculate available days for each user based on manager templates."""
        if not hasattr(self, "manager_templates"):
            return

        for user in self.user_pw_day_usage.keys():
            # Find which manager template applies to this user
            applicable_weeks = []

            for manager_id, pattern in self.user_type_mapping.items():
                if re.match(pattern, user):
                    applicable_weeks = self.manager_templates.get(manager_id, [])
                    break

            # Generate all PW-D combinations for applicable weeks
            if applicable_weeks:
                for week in applicable_weeks:
                    for day in range(1, 8):  # D1 to D7
                        pw_day = f"{week}-D{day}"
                        self.user_available_days[user].add(pw_day)

        print(f"Calculated available days for {len(self.user_available_days)} users")

    def parse_planning_week_day(self, pw_day_string, planning_week_start=None):
        """
        Parse planning week day format like 'PW1-D1', 'PW2-D3', etc.
        Returns tuple: (actual_date, pw_string)
        """
        if not pw_day_string:
            return None, None

        match = re.match(r"PW(\d+)-D(\d+)", str(pw_day_string).strip())
        if not match:
            return None, None

        week = int(match.group(1))
        day = int(match.group(2))

        # Calculate offset from planning week start
        day_offset = (week - 1) * 7 + (day - 1)

        if planning_week_start:
            try:
                if isinstance(planning_week_start, str):
                    base_date = datetime.strptime(planning_week_start, "%Y%m%d").date()
                else:
                    base_date = planning_week_start
            except:
                base_date = datetime.now().date()
        else:
            base_date = datetime.now().date()

        target_date = base_date + timedelta(days=day_offset)
        return target_date, pw_day_string

    def extract_users(self, value):
        """
        Extract users from a value string.
        """
        users = []
        if isinstance(value, str):
            patterns = [
                r"SIT-ESS-[A-Z]+-USER\d+",
                r"SIT-SM-[A-Z]+-USER\d+",
                r"QA-ESS-[A-Z0-9\-]+",
                r"Store Admin\d+",
                r"SYSADMIN-USER\d+",
            ]

            for pattern in patterns:
                found_users = re.findall(pattern, value, re.IGNORECASE)
                users.extend(found_users)

        return list(set(users))

    def get_date_range(self, start_date, end_date):
        """
        Get all dates between start_date and end_date (inclusive)
        """
        if not start_date:
            return []

        if not end_date:
            return [start_date]

        date_range = []
        current_date = start_date
        while current_date <= end_date:
            date_range.append(current_date)
            current_date += timedelta(days=1)

        return date_range

    def load_and_parse_csv(self):
        """
        Load the CSV file and parse all test data
        """
        planning_week_start = None

        try:
            with open(self.csv_file_path, "r", encoding="utf-8") as file:
                reader = csv.DictReader(file)

                # First pass: collect all data and find planning week start
                for row in reader:
                    self.test_data.append(row)

                    if row.get("KEY") == "PLANNING_WEEK_START":
                        planning_week_start = row.get("VALUE", "").strip()

                # Second pass: analyze test cases
                test_case_dates = {}  # {test_id: {key: (date, pw_string)}}
                test_case_users = {}  # {test_id: [users]}

                for row in self.test_data:
                    scope = row.get("SCOPE", "UNKNOWN").strip()

                    if scope == "GLOBAL":
                        continue

                    test_id = scope
                    key = row.get("KEY", "").strip()
                    value = row.get("VALUE", "").strip()

                    self.test_case_info[test_id].append(row)

                    # Extract users from ESS_USER* keys (ESS_USER, ESS_USER1, ESS_USER2, ESS_USER3, ESS_USER_NEARBY, etc.)
                    # Using pattern matching to catch all variations
                    if key.startswith("ESS_USER"):
                        if test_id not in test_case_users:
                            test_case_users[test_id] = []
                        users = self.extract_users(value)
                        test_case_users[test_id].extend(users)

                    # Extract dates from fields that contain PW-D format
                    # Exclude configuration fields that define test ranges (not actual action dates)
                    excluded_date_fields = [
                        "START_DATE",
                        "END_DATE",
                        "START_WEEK_DATE",
                        "END_WEEK_DATE",
                        "AVAIL_WEEK_START_DATE",
                        "AVAIL_WEEK_END_DATE",
                        "REQ_START_WEEK_DATE",
                        "REQ_APPROVED_START_WEEK_DATE",
                        "REQ_APPROVED_END_WEEK_DATE",
                        "REQ_DECLINED_START_WEEK_DATE",
                        "REQ_DECLINED_END_WEEK_DATE",
                        "REQ_PENDING_START_WEEK_DATE",
                        "REQ_PENDING_END_WEEK_DATE",
                        "APPROVED_REQ_START_DATE",
                        "APPROVED_REQ_END_DATE",
                        "PENDING_REQ_START_DATE",
                        "PENDING_REQ_END_DATE",
                        "CANCELLED_REQ_START_DATE",
                        "CANCELLED_REQ_END_DATE",
                        "DECLINED_REQ_START_DATE",
                        "REQ_START_DATE",
                        "PERM_APPROVED_AVAIL_WEEK_START_DATE",
                        "PERM_DECLINED_AVAIL_WEEK_START_DATE",
                        "PERM_PENDING_AVAIL_WEEK_START_DATE",
                        "TEMP_APPROVED_AVAIL_WEEK_START_DATE",
                        "TEMP_APPROVED_AVAIL_WEEK_END_DATE",
                        "TEMP_DECLINED_AVAIL_WEEK_START_DATE",
                        "TEMP_DECLINED_AVAIL_WEEK_END_DATE",
                        "TEMP_PENDING_AVAIL_WEEK_START_DATE",
                        "TEMP_PENDING_AVAIL_WEEK_END_DATE",
                        "REQ_PARTIAL_APPROVED_START_WEEK_DATE",
                        # Week start/end configuration fields (not action dates)
                        "PUBLISHED_WEEK_START_DATE",
                        "UNPUBLISHED_WEEK_START_DATE",
                        "TIME_CARD_PLANNING_WEEK_START_DATE",
                    ]

                    # If field contains PW-D format AND is not an excluded config field, capture it
                    if (
                        re.match(r"PW\d+-D\d+", value)
                        and key not in excluded_date_fields
                    ):
                        date, pw_string = self.parse_planning_week_day(
                            value, planning_week_start
                        )
                        if date:
                            if test_id not in test_case_dates:
                                test_case_dates[test_id] = {}
                            test_case_dates[test_id][key] = (date, pw_string)
                            self.date_to_pw_map[date] = pw_string

                # Third pass: match users with dates for each test case
                for test_id in self.test_case_info.keys():
                    users = test_case_users.get(test_id, [])
                    date_dict = test_case_dates.get(test_id, {})

                    if users:
                        self.user_usage_by_test[test_id].update(users)

                    # Collect all PW-D strings for this test (only from specific date fields)
                    all_pw_strings = set()

                    # Add all captured dates (no range expansion) and normalize format
                    for key, (date, pw_string) in date_dict.items():
                        normalized_pw = self.normalize_pw_format(pw_string)
                        all_pw_strings.add(normalized_pw)

                    # Track user+PW combinations with test case IDs
                    for user in users:
                        for pw_string in all_pw_strings:
                            self.user_pw_day_usage[user].add(pw_string)

                            # Track which test case uses this day for this user
                            self.user_day_test_mapping[(user, pw_string)].append(
                                test_id
                            )

            print(f"Successfully loaded {len(self.test_data)} rows from CSV")
            print(f"Planning Week Start: {planning_week_start}")
            print(f"Identified {len(self.test_case_info)} test cases")
            print(f"Identified {len(self.user_pw_day_usage)} unique users")
            return True

        except FileNotFoundError:
            print(f"Error: File not found: {self.csv_file_path}")
            return False
        except Exception as e:
            print(f"Error reading CSV: {e}")
            import traceback

            traceback.print_exc()
            return False

    def generate_excel_report(self, output_file):
        """
        Generate an Excel report with user details and day usage
        """
        if not EXCEL_AVAILABLE:
            print("Excel generation not available. Generating CSV instead.")
            self.generate_csv_report(output_file.replace(".xlsx", ".csv"))
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "User Day Usage"

        # Define styles
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Simplified headers - removed duplicate column as it's redundant
        headers = [
            "User Name",
            "Days Used (PW-D with Test Cases)",
            "Days Available",
            "Total Days Used",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = border

        # Set column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 60  # Days used with test cases
        ws.column_dimensions["C"].width = 50  # Days available
        ws.column_dimensions["D"].width = 15  # Total count

        # Sort users alphabetically
        sorted_users = sorted(self.user_pw_day_usage.keys())

        # Add data rows
        current_row = 2
        for user in sorted_users:
            # Get all days used for this user, sorted by PW and day number
            all_days = sorted(
                self.user_pw_day_usage.get(user, set()),
                key=lambda x: (
                    int(re.search(r"PW(\d+)", x).group(1)),
                    int(re.search(r"D(\d+)", x).group(1)),
                ),
            )

            # Build days used string with test case IDs
            days_with_tests = []
            for pw_day in all_days:
                test_cases = self.user_day_test_mapping.get((user, pw_day), [])
                test_list = ", ".join(sorted(set(test_cases)))
                days_with_tests.append(f"{pw_day} ({test_list})")

            days_used_str = "\n".join(days_with_tests) if days_with_tests else "-"
            total_days = len(all_days)

            # Add user name
            cell = ws.cell(row=current_row, column=1, value=user)
            cell.alignment = Alignment(vertical="top")
            cell.border = border

            # Add days used with test cases
            cell = ws.cell(row=current_row, column=2, value=days_used_str)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

            # Calculate available days - remove ALL used days
            all_available = self.user_available_days.get(user, set())
            all_used = self.user_pw_day_usage.get(user, set())
            available_diff = all_available - all_used
            truly_available = (
                sorted(
                    available_diff,
                    key=lambda x: (
                        int(re.search(r"PW(\d+)", x).group(1)),
                        int(re.search(r"D(\d+)", x).group(1)),
                    ),
                )
                if available_diff
                else []
            )
            available_str = ", ".join(truly_available) if truly_available else "-"

            # Add days available
            cell = ws.cell(row=current_row, column=3, value=available_str)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            cell.fill = PatternFill(
                start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
            )

            # Add total
            cell = ws.cell(row=current_row, column=4, value=total_days)
            cell.alignment = Alignment(vertical="top")
            cell.border = border

            current_row += 1

        # Freeze the header row
        ws.freeze_panes = "A2"

        # Save the workbook
        wb.save(output_file)
        print(f"\nExcel report generated: {output_file}")

    def generate_csv_report(self, output_file):
        """
        Generate a CSV report with user details and day usage
        """
        with open(output_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)

            # Write headers
            writer.writerow(
                [
                    "User Name",
                    "Days Used (PW-D with Test Cases)",
                    "Days Available",
                    "Total Days Used",
                ]
            )

            # Sort users alphabetically
            sorted_users = sorted(self.user_pw_day_usage.keys())

            # Write data rows
            for user in sorted_users:
                # Get all days used for this user
                all_days = sorted(
                    self.user_pw_day_usage.get(user, set()),
                    key=lambda x: (
                        int(re.search(r"PW(\d+)", x).group(1)),
                        int(re.search(r"D(\d+)", x).group(1)),
                    ),
                )

                # Build days used string with test case IDs
                days_with_tests = []
                for pw_day in all_days:
                    test_cases = self.user_day_test_mapping.get((user, pw_day), [])
                    test_list = ", ".join(sorted(set(test_cases)))
                    days_with_tests.append(f"{pw_day} ({test_list})")

                days_used_str = ", ".join(days_with_tests) if days_with_tests else "-"

                # Calculate available days
                all_available = self.user_available_days.get(user, set())
                all_used = self.user_pw_day_usage.get(user, set())
                available_diff = all_available - all_used
                truly_available = (
                    sorted(
                        available_diff,
                        key=lambda x: (
                            int(re.search(r"PW(\d+)", x).group(1)),
                            int(re.search(r"D(\d+)", x).group(1)),
                        ),
                    )
                    if available_diff
                    else []
                )
                available_str = ", ".join(truly_available) if truly_available else "-"

                total_days = len(all_days)

                writer.writerow([user, days_used_str, available_str, total_days])

        print(f"\nCSV report generated: {output_file}")

    def analyze(self, output_file=None):
        """
        Perform analysis and generate report

        Args:
            output_file: Path to output file. If None, auto-generates based on input CSV path
        """
        if not self.load_and_parse_csv():
            return

        # Calculate available days based on manager templates
        self.calculate_available_days()

        print("\n" + "=" * 80)
        print("TEST DATA ANALYSIS REPORT")
        print("=" * 80)
        print(f"Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"CSV File: {self.csv_file_path}")
        print(f"Total Test Cases: {len(self.test_case_info)}")
        print(f"Total Data Rows: {len(self.test_data)}")
        print("=" * 80)

        # Generate output file path if not provided
        if output_file is None:
            # Try to generate in same directory as CSV
            csv_dir = os.path.dirname(self.csv_file_path)
            output_file = os.path.join(csv_dir, "user_day_usage_report.xlsx")
        if EXCEL_AVAILABLE:
            self.generate_excel_report(output_file)
        else:
            csv_output = output_file.replace(".xlsx", ".csv")
            self.generate_csv_report(csv_output)

        # Print summary
        print("\n" + "-" * 80)
        print("USER USAGE SUMMARY")
        print("-" * 80)
        print(f"Total Unique Users: {len(self.user_pw_day_usage)}")

        for user in sorted(self.user_pw_day_usage.keys())[:10]:  # Show first 10
            pw_days = sorted(
                self.user_pw_day_usage[user],
                key=lambda x: (
                    int(re.search(r"PW(\d+)", x).group(1)),
                    int(re.search(r"D(\d+)", x).group(1)),
                ),
            )

            print(f"\n{user}:")
            print(f"  Total: {len(pw_days)} days")
            if pw_days:
                # Show first 5 days with their test cases
                sample_days = pw_days[:5]
                for pw_day in sample_days:
                    test_cases = self.user_day_test_mapping.get((user, pw_day), [])
                    test_list = ", ".join(sorted(set(test_cases)))
                    print(f"  {pw_day} (used in: {test_list})")
                if len(pw_days) > 5:
                    print(f"  ... and {len(pw_days) - 5} more days")

        if len(self.user_pw_day_usage) > 10:
            print(f"\n... and {len(self.user_pw_day_usage) - 10} more users")

        # Report duplicate days (reused in multiple tests)
        print("\n" + "=" * 80)
        print("DUPLICATE DAY USAGE DETECTION")
        print("=" * 80)

        duplicates_found = False
        for user in sorted(self.user_pw_day_usage.keys()):
            user_duplicates = []
            for pw_day in sorted(
                self.user_pw_day_usage[user],
                key=lambda x: (
                    int(re.search(r"PW(\d+)", x).group(1)),
                    int(re.search(r"D(\d+)", x).group(1)),
                ),
            ):
                test_cases = self.user_day_test_mapping.get((user, pw_day), [])
                if len(test_cases) > 1:
                    test_list = ", ".join(sorted(set(test_cases)))
                    user_duplicates.append(f"  ⚠️  {pw_day} is reused in: {test_list}")

            if user_duplicates:
                duplicates_found = True
                print(f"\n{user}:")
                for dup in user_duplicates:
                    print(dup)

        if not duplicates_found:
            print("\n✅ No duplicate day usage detected! All days are unique per user.")
        else:
            print("\n⚠️  WARNING: Found duplicate day usage across multiple test cases!")
            print(
                "   This means the same day is being used for the same user in different tests."
            )
            print("   Consider using different days to avoid conflicts.")

        print("\n" + "=" * 80)
        print("Report generated successfully!")
        print("=" * 80)


def main():
    """
    Main function with command-line argument support
    """
    parser = argparse.ArgumentParser(
        description="Analyze test data CSV and generate user/day usage report",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Use default files (VARIATION1.csv)
  python3 analyze_testdata.py
  
  # Specify custom CSV file (auto-detect output name)
  python3 analyze_testdata.py TestData/VARIATION2.csv
  
  # Specify CSV and template files
  python3 analyze_testdata.py TestData/VARIATION2.csv TestData/custom_template.json
  
  # Specify all files including output
  python3 analyze_testdata.py TestData/VARIATION2.csv TestData/custom_template.json output/variation2_report.xlsx
        """,
    )

    parser.add_argument(
        "csv_file",
        nargs="?",
        help="Path to the CSV file to analyze (default: ../../../TestData/VARIATION1.csv)",
    )
    parser.add_argument(
        "template_file",
        nargs="?",
        help="Path to the template mapping JSON file (default: ../../../TestData/template_mapping_shift_data/template_week_mapping.json)",
    )
    parser.add_argument(
        "output_file",
        nargs="?",
        help="Path to the output Excel file (default: reports/user_day_usage_report.xlsx)",
    )

    args = parser.parse_args()

    # Get script directory and calculate project root
    script_dir = Path(__file__).parent.resolve()
    project_root = (
        script_dir.parent.parent.parent
    )  # Library/Mobile/test_data_analyzer -> rflx-wfm-sanity

    # Set default paths (relative to project root)
    default_csv = project_root / "TestData" / "VARIATION1.csv"
    default_template = (
        project_root
        / "TestData"
        / "template_mapping_shift_data"
        / "template_week_mapping.json"
    )

    # Determine file paths
    if args.csv_file:
        csv_file = Path(args.csv_file)
        if not csv_file.is_absolute():
            # Try current working directory first, then script directory
            cwd_path = Path.cwd() / csv_file
            if cwd_path.exists():
                csv_file = cwd_path
            else:
                csv_file = script_dir / csv_file
    else:
        csv_file = default_csv

    if args.template_file:
        template_file = Path(args.template_file)
        if not template_file.is_absolute():
            # Try current working directory first, then script directory
            cwd_path = Path.cwd() / template_file
            if cwd_path.exists():
                template_file = cwd_path
            else:
                template_file = script_dir / template_file
    else:
        template_file = default_template

    if args.output_file:
        output_file = Path(args.output_file)
        if not output_file.is_absolute():
            output_file = script_dir / output_file
    else:
        # Auto-generate output filename based on input CSV name (variation name as prefix)
        csv_name = csv_file.stem.lower()  # Get variation name from CSV file
        output_file = script_dir / "reports" / f"{csv_name}_report.xlsx"

    # Validate input files exist
    if not csv_file.exists():
        print(f"ERROR: CSV file not found: {csv_file}")
        print("Please check the path and try again.")
        sys.exit(1)

    if not template_file.exists():
        print(f"WARNING: Template file not found: {template_file}")
        print("Available days calculation will be skipped.")
        use_template = False
    else:
        use_template = True

    # Create output directory if it doesn't exist
    output_file.parent.mkdir(parents=True, exist_ok=True)

    # Print configuration
    print("=" * 80)
    print("TEST DATA ANALYZER - CONFIGURATION")
    print("=" * 80)
    print(f"CSV Input File:     {csv_file}")
    print(f"Template File:      {template_file if use_template else 'Not used'}")
    print(f"Output File:        {output_file}")
    print("=" * 80)
    print()

    # Run analysis
    analyzer = TestDataAnalyzer(str(csv_file))
    if use_template:
        analyzer.load_template_mapping(str(template_file))
    analyzer.analyze(str(output_file))


if __name__ == "__main__":
    main()
